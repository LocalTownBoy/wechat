import datetime
import io
import json
import logging
import os
import re
import time

from django.conf import settings
from django.http import JsonResponse
from django.shortcuts import render
from openpyxl import Workbook, load_workbook
import requests
import pdfplumber
from qcloud_cos import CosConfig, CosS3Client  # COS SDK，用于上传 PDF
from wxcloudrun.models import Counters, Paper


logger = logging.getLogger('log')
EXCEL_FILE_NAME = 'push_messages.xlsx'
WX_APPID = os.getenv("WX_APPID")
WX_SECRET = os.getenv("WX_SECRET")
_token_cache = {"value": None, "expires_at": 0}
# COS 配置（论文 PDF 会上传到对象存储）
COS_SECRET_ID = os.getenv("COS_SECRET_ID_pdf")
COS_SECRET_KEY = os.getenv("COS_SECRET_KEY_pdf")
COS_REGION = os.getenv("COS_REGION_pdf")
COS_BUCKET = os.getenv("COS_BUCKET_pdf")
COS_PREFIX = os.getenv("COS_PREFIX", "papers/")


def index(request, _):
    """
    获取主页

     `` request `` 请求对象
    """

    return render(request, 'index.html')


def counter(request, _):
    
    """
    获取当前计数

     `` request `` 请求对象
    """

    rsp = JsonResponse({'code': 0, 'errorMsg': ''}, json_dumps_params={'ensure_ascii': False})
    if request.method == 'GET' or request.method == 'get':
        rsp = get_count()
    elif request.method == 'POST' or request.method == 'post':
        rsp = update_count(request)
    else:
        rsp = JsonResponse({'code': -1, 'errorMsg': '请求方式错误'},
                            json_dumps_params={'ensure_ascii': False})
    logger.info('response result: {}'.format(rsp.content.decode('utf-8')))
    return rsp


def push_msg(request, _):
    """
    接收微信公众号推送的JSON消息并写入Excel
    """

    if request.method not in ['POST', 'post']:
        return JsonResponse({'code': -1, 'errorMsg': '请求方式错误'},
                            json_dumps_params={'ensure_ascii': False}, status=405)

    if not request.body:
        return JsonResponse({'code': -1, 'errorMsg': '请求体为空'},
                            json_dumps_params={'ensure_ascii': False})

    try:
        body_unicode = request.body.decode('utf-8')
        payload = json.loads(body_unicode)
    except json.JSONDecodeError:
        return JsonResponse({'code': -1, 'errorMsg': '请求体不是合法JSON'},
                            json_dumps_params={'ensure_ascii': False})

    try:
        file_path = _append_to_excel(payload)
    except Exception as exc:  # pylint: disable=broad-except
        logger.exception('failed to write push message to excel: %s', exc)
        return JsonResponse({'code': -1, 'errorMsg': '写入Excel失败'},
                            json_dumps_params={'ensure_ascii': False})

    return JsonResponse({'code': 0, 'data': {'file': file_path}},
                        json_dumps_params={'ensure_ascii': False})


def push_msg_list(request, _):
    """
    展示Excel中已存储的推送消息
    """
    excel_path = getattr(settings, 'PUSH_MSG_EXCEL_PATH',
                         os.path.join(settings.BASE_DIR, EXCEL_FILE_NAME))
    messages = []
    error = None

    if not os.path.exists(excel_path):
        error = 'Excel文件不存在，请先调用 /push/msg 写入数据'
    else:
        try:
            workbook = load_workbook(excel_path)
            sheet = workbook.active
            for received_at, payload in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                if received_at is None and payload is None:
                    continue
                payload_text = payload
                if isinstance(payload, str):
                    try:
                        payload_text = json.dumps(json.loads(payload), ensure_ascii=False, indent=2)
                    except json.JSONDecodeError:
                        payload_text = payload
                messages.append({'received_at': received_at, 'payload': payload_text})
        except Exception as exc:  # pylint: disable=broad-except
            logger.exception('failed to load excel: %s', exc)
            error = '读取Excel失败，请检查日志'

    return render(request, 'push_messages.html', {
        'messages': messages,
        'excel_path': excel_path,
        'error': error,
    })


def papers(request, _):
    """
    收集论文信息（标题、作者、章节）并存入数据库，返回列表页面
    - 支持直接填写文本，也支持上传 PDF 自动解析并上传 COS
    """
    message = None
    error = None
    parsed = {}

    if request.method in ['POST', 'post']:
        title = request.POST.get('title', '').strip()
        author = request.POST.get('author', '').strip()
        section = request.POST.get('section', '').strip()
        pdf_file = request.FILES.get('pdf')
        upload_url = None
        pdf_bytes = None

        if pdf_file:
            try:
                pdf_bytes = pdf_file.read()  # 原始文件内容
                parsed = _parse_pdf(io.BytesIO(pdf_bytes), getattr(pdf_file, 'name', ''))
                title = title or parsed.get('title', '')
                author = author or parsed.get('author', '')
                section = section or parsed.get('sections_text', '')
                upload_url = _upload_pdf_to_cos(pdf_bytes, getattr(pdf_file, 'name', 'paper.pdf'))
            except Exception as exc:  # pylint: disable=broad-except
                logger.exception('failed to parse pdf: %s', exc)
                error = 'PDF 解析失败，请检查文件格式'

        if not title or not author or not section:
            error = '请完整填写标题、作者和章节'
        else:
            try:
                # 保存解析/用户填写的信息及文件 URL
                Paper.objects.create(title=title, author=author, section=section, url=upload_url)
                message = '已保存！'
            except Exception as exc:  # pylint: disable=broad-except
                logger.exception('failed to save paper: %s', exc)
                error = '保存失败，请检查日志'

    try:
        # 从数据库中获取论文列表
        papers = Paper.objects.order_by('-id').all()
    except Exception as exc:  # pylint: disable=broad-except
        logger.exception('failed to load papers: %s', exc)
        papers = []
        if not error:
            error = '读取数据失败，请检查日志'

    return render(request, 'paper_form.html', {
        'papers': papers,
        'message': message,
        'error': error,
        'parsed': parsed,
    })


def wx_send_message(request, _):
    """
    主动向公众号用户推送消息（基于微信群发接口）
    请求示例:
    {
        "touser": ["openid1", "openid2"],
        "msgtype": "text",
        "text": {"content": "你好"},
        "send_ignore_reprint": 0
    }
    """
    if request.method not in ['POST', 'post']:
        return JsonResponse({'code': -1, 'errorMsg': '请求方式错误'},
                            json_dumps_params={'ensure_ascii': False}, status=405)

    try:
        body = json.loads(request.body.decode('utf-8'))
    except Exception:  # pylint: disable=broad-except
        return JsonResponse({'code': -1, 'errorMsg': '请求体需为JSON'},
                            json_dumps_params={'ensure_ascii': False})

    touser = body.get('touser') or []
    msgtype = (body.get('msgtype') or '').strip()
    send_ignore_reprint = body.get('send_ignore_reprint')
    text_obj = body.get('text') or {}

    if len(touser) > 10000:
        touser = touser[:10000]

    if msgtype not in ['text']:
        return JsonResponse({'code': -1, 'errorMsg': '当前仅支持msgtype=text'},
                            json_dumps_params={'ensure_ascii': False})

    if msgtype == 'text':
        content = text_obj.get('content', '').strip()
        if not content:
            return JsonResponse({'code': -1, 'errorMsg': 'text.content 不能为空'},
                                json_dumps_params={'ensure_ascii': False})

    if not WX_APPID or not WX_SECRET:
        return JsonResponse({'code': -1, 'errorMsg': '未配置WX_APPID/WX_SECRET'},
                            json_dumps_params={'ensure_ascii': False})

    payload = {
        "touser": touser,
        "msgtype": msgtype,
    }
    if send_ignore_reprint is not None:
        payload["send_ignore_reprint"] = send_ignore_reprint

    if msgtype == 'text':
        payload["text"] = {"content": content}

    try:
        # token = _get_access_token()
        data=_send_mass_message(payload)
    except Exception as exc:  # pylint: disable=broad-except
        logger.exception('failed to send wechat message: %s', exc)
        return JsonResponse({'code': -1, 'errorMsg': '发送失败，请查看日志'},
                            json_dumps_params={'ensure_ascii': False})

    return JsonResponse({'code': 0, 'data': data}, json_dumps_params={'ensure_ascii': False})


def wx_users_info(request, _):
    """
    获取关注用户基本信息（批量，每次最多100个 openid）
    """
    if request.method not in ['POST', 'post']:
        return JsonResponse({'code': -1, 'errorMsg': '请求方式错误'},
                            json_dumps_params={'ensure_ascii': False}, status=405)

    try:
        body = json.loads(request.body.decode('utf-8'))
    except Exception:  # pylint: disable=broad-except
        return JsonResponse({'code': -1, 'errorMsg': '请求体需为JSON'},
                            json_dumps_params={'ensure_ascii': False})

    openids = body.get('openids') or []
    lang = body.get('lang', 'zh_CN')

    if not isinstance(openids, list) or not openids:
        return JsonResponse({'code': -1, 'errorMsg': 'openids需为非空数组'},
                            json_dumps_params={'ensure_ascii': False})

    if len(openids) > 100:
        openids = openids[:100]

    if not WX_APPID or not WX_SECRET:
        return JsonResponse({'code': -1, 'errorMsg': '未配置WX_APPID/WX_SECRET'},
                            json_dumps_params={'ensure_ascii': False})

    try:
        token = _get_access_token()
        data = _batch_get_users(token, openids, lang)
    except Exception as exc:  # pylint: disable=broad-except
        logger.exception('failed to get users info: %s', exc)
        return JsonResponse({'code': -1, 'errorMsg': '获取失败，请查看日志'},
                            json_dumps_params={'ensure_ascii': False})

    return JsonResponse({'code': 0, 'data': data},
                        json_dumps_params={'ensure_ascii': False})


def get_count():
    """
    获取当前计数
    """

    try:
        data = Counters.objects.get(id=1)
    except Counters.DoesNotExist:
        return JsonResponse({'code': 0, 'data': 0},
                    json_dumps_params={'ensure_ascii': False})
    return JsonResponse({'code': 0, 'data': data.count},
                        json_dumps_params={'ensure_ascii': False})


def update_count(request):
    """
    更新计数，自增或者清零

    `` request `` 请求对象
    """

    logger.info('update_count req: {}'.format(request.body))

    body_unicode = request.body.decode('utf-8')
    body = json.loads(body_unicode)

    if 'action' not in body:
        return JsonResponse({'code': -1, 'errorMsg': '缺少action参数'},
                            json_dumps_params={'ensure_ascii': False})

    if body['action'] == 'inc':
        try:
            data = Counters.objects.get(id=1)
        except Counters.DoesNotExist:
            data = Counters()
        data.id = 1
        data.count += 1
        data.save()
        return JsonResponse({'code': 0, "data": data.count},
                    json_dumps_params={'ensure_ascii': False})
    elif body['action'] == 'clear':
        try:
            data = Counters.objects.get(id=1)
            data.delete()
        except Counters.DoesNotExist:
            logger.info('record not exist')
        return JsonResponse({'code': 0, 'data': 0},
                    json_dumps_params={'ensure_ascii': False})
    else:
        return JsonResponse({'code': -1, 'errorMsg': 'action参数错误'},
                    json_dumps_params={'ensure_ascii': False})


def _append_to_excel(payload):
    """
    将推送消息写入Excel，按时间顺序追加
    """
    # 获取Excel文件路径
    excel_path = getattr(settings, 'PUSH_MSG_EXCEL_PATH',
                         os.path.join(settings.BASE_DIR, EXCEL_FILE_NAME))
    directory = os.path.dirname(excel_path)

    if directory and not os.path.exists(directory):
        os.makedirs(directory, exist_ok=True)
    # 创建文件夹
    workbook = _load_or_create_workbook(excel_path)
    # 获取活动表
    sheet = workbook.active
    # 追加数据行
    sheet.append([
        datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        json.dumps(payload, ensure_ascii=False)
    ])
    # 保存文件
    workbook.save(excel_path)
    return excel_path


def _load_or_create_workbook(excel_path):
    """
    获取已存在的工作簿，或创建新的并初始化表头
    """
    if os.path.exists(excel_path):
        return load_workbook(excel_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'messages'
    sheet.append(['received_at', 'payload'])
    return workbook


def _parse_pdf(pdf_file, filename=''):
    """
    解析上传的 PDF，返回标题、作者、章节列表的最佳猜测
    - 优先读取元数据，其次读取前几页文本，最后回退文件名
    """
    try:
        pdf_file.seek(0)
    except Exception:  # pylint: disable=broad-except
        pass

    fname = filename or getattr(pdf_file, 'name', '') or ''
    base_title = os.path.splitext(os.path.basename(fname))[0] if fname else ''

    with pdfplumber.open(pdf_file) as pdf:
        # 尝试读取元数据
        info = pdf.metadata or {}
        title = ''
        author = ''

        raw_title = info.get('Title')
        raw_author = info.get('Author')
        if raw_title:
            title = str(raw_title).strip()
        if raw_author:
            author = str(raw_author).strip()

        # 抽取前几页文本用于解析章节/备用标题
        texts = []
        for page in pdf.pages[:5]:
            page_text = page.extract_text() or ''
            texts.append(page_text)
        full_text = '\n'.join(texts)

    lines = [ln.strip() for ln in full_text.splitlines() if ln.strip()]

    if not title:
        if lines:
            title = lines[0][:255]
        elif base_title:
            title = base_title[:255]

    if not author:
        for ln in lines[:5]:
            m = re.search(r'(作者|Author)[:：]\s*(.+)', ln)
            if m:
                author = m.group(2).strip()[:255]
                break

    sections = []
    section_pattern = re.compile(r'^(\d+(\.\d+)*)\s+(.+)|^(摘要|Abstract|引言|绪论|结论|参考文献)', re.IGNORECASE)
    for ln in lines:
        if len(sections) >= 15:
            break
        if section_pattern.match(ln):
            sections.append(ln[:255])

    dedup_sections = []
    for sec in sections:
        if sec not in dedup_sections:
            dedup_sections.append(sec)

    sections_text = '; '.join(dedup_sections[:10])

    return {
        'title': title,
        'author': author,
        'sections': dedup_sections,
        'sections_text': sections_text,
    }


def _get_access_token():
    """
    获取并缓存公众号 access_token
    """
    now = int(time.time())
    if _token_cache["value"] and _token_cache["expires_at"] > now + 60:
        return _token_cache["value"]

    url = "https://api.weixin.qq.com/cgi-bin/token"
    resp = requests.get(url, params={"grant_type": "client_credential",
                                     "appid": WX_APPID,
                                     "secret": WX_SECRET}, timeout=5)
    logger.info("get access_token resp: %s", resp.text)
    data = resp.json()
    if "access_token" not in data:
        raise RuntimeError(f"get access_token failed: {data}")
    _token_cache["value"] = data["access_token"]
    _token_cache["expires_at"] = now + int(data.get("expires_in", 7200))
    return _token_cache["value"]



def _batch_get_users(token, openids, lang):
    """
    调用批量获取用户基本信息接口
    """
    url = "https://api.weixin.qq.com/cgi-bin/user/info/batchget"
    payload = {
        "user_list": [{"openid": oid, "lang": lang} for oid in openids]
    }
    resp = requests.post(url, params={"access_token": token},
                         json=payload, timeout=5)
    data = resp.json()
    if data.get("errcode", 0) not in (0, None):
        raise RuntimeError(f"batchget failed: {data}")
    return data.get("user_info_list", [])


def _send_mass_message(payload):
    """
    群发接口（按 openid 列表）
    """
    url = "http://api.weixin.qq.com/cgi-bin/message/mass/send"
    resp = requests.post(url,
                         json=payload, timeout=5)
    data = resp.json()
    if data.get("errcode", 0) != 0:
        raise RuntimeError(f"mass send failed: {data}")
    return data


def _upload_pdf_to_cos(pdf_bytes, filename):
    """
    上传 PDF 到腾讯云 COS，返回可访问的 URL
    - 依赖环境变量 COS_SECRET_ID/KEY/REGION/BUCKET
    - COS_PREFIX 可选，默认 papers/
    """
    if not (COS_SECRET_ID and COS_SECRET_KEY and COS_REGION and COS_BUCKET):
        raise RuntimeError('未配置 COS_SECRET_ID/COS_SECRET_KEY/COS_REGION/COS_BUCKET')

    config = CosConfig(Region=COS_REGION, SecretId=COS_SECRET_ID, SecretKey=COS_SECRET_KEY)
    client = CosS3Client(config)

    safe_prefix = COS_PREFIX if COS_PREFIX.endswith('/') else COS_PREFIX + '/'
    key = f"{safe_prefix}{int(time.time())}_{os.path.basename(filename or 'paper.pdf')}"

    client.put_object(
        Bucket=COS_BUCKET,
        Body=pdf_bytes,
        Key=key,
        ContentType='application/pdf'
    )
    url = f"https://{COS_BUCKET}.cos.{COS_REGION}.myqcloud.com/{key}"
    return url
